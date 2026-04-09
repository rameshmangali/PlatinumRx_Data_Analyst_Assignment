import sys
import subprocess

def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("openpyxl not found. Installing...")
    install("openpyxl")
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill

import os

def generate_excel():
    # Setup directory
    folder_path = r"e:\PROJECT\PlatinumRx\Data_Analyst_Assignment\Spreadsheets"
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    file_path = os.path.join(folder_path, "Ticket_Analysis.xlsx")
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # --- 1. TICKET SHEET ---
    ws_ticket = wb.active
    ws_ticket.title = "ticket"
    
    # Headers
    ticket_headers = ["ticket_id", "created_at", "closed_at", "outlet_id", "cms_id"]
    ws_ticket.append(ticket_headers)
    
    # Data
    ticket_data = [
        ["isu-sjd-457", "2021-08-19 16:45:43", "2021-08-22 12:33:32", "wrqy-juv-978", "vew-iuvd-12"],
        ["qer-fal-092", "2021-08-21 11:09:22", "2021-08-21 17:13:45", "8woh-k3u-23b", "cms-test-01"],
        ["sam-ple-001", "2021-08-21 11:15:00", "2021-08-21 11:45:00", "8woh-k3u-23b", "cms-test-02"], # Same day, same hour
        ["sam-ple-002", "2021-08-22 09:00:00", "2021-08-22 20:00:00", "wrqy-juv-978", "cms-test-03"], # Same day, diff hour
    ]
    for row in ticket_data:
        ws_ticket.append(row)
        
    # --- Helper columns for Question 2 in ticket sheet ---
    # F: Same Day?
    # G: Same Hour?
    ws_ticket["F1"] = "Same Day?"
    ws_ticket["G1"] = "Same Hour?"
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for cell in ws_ticket[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    for row_idx in range(2, len(ticket_data) + 2):
        # We need to extract the date portion. In Excel, we can use INT() if it's a date number, 
        # but since we have text strings like '2021-08-19 16:45:43', we can use LEFT(text, 10).
        # Same day check: =LEFT(B2, 10) = LEFT(C2, 10)
        ws_ticket[f"F{row_idx}"] = f"=LEFT(B{row_idx}, 10)=LEFT(C{row_idx}, 10)"
        
        # Same hour check: same day AND same hour (characters 12-13)
        # =AND(F2, MID(B2, 12, 2)=MID(C2, 12, 2))
        ws_ticket[f"G{row_idx}"] = f"=AND(F{row_idx}, MID(B{row_idx}, 12, 2)=MID(C{row_idx}, 12, 2))"
    
    # --- 2. FEEDBACKS SHEET ---
    ws_feedback = wb.create_sheet(title="feedbacks")
    
    feedback_headers = ["cms_id", "feedback_at", "feedback_rating", "ticket_created_at"]
    ws_feedback.append(feedback_headers)
    
    for cell in ws_feedback[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    feedback_data = [
        ["vew-iuvd-12", "2021-08-21 13:26:48", 3],
        ["cms-test-01", "2021-08-22 10:00:00", 4],
        ["cms-test-missing", "2021-08-23 10:00:00", 5],
    ]
    for row in feedback_data:
        ws_feedback.append(row)
        
    # Apply VLOOKUP for Question 1
    # We want to pull 'created_at' from 'ticket' sheet where cms_id matches.
    # cms_id is column E (5th column) in ticket sheet. 'created_at' is column B.
    # Standard VLOOKUP only works left-to-right (lookup value must be in first column of range).
    # Since cms_id is in E and created_at in B, we must use INDEX-MATCH.
    # =INDEX(ticket!B:B, MATCH(A2, ticket!E:E, 0))
    for row_idx in range(2, len(feedback_data) + 2):
        ws_feedback[f"D{row_idx}"] = f"=IFERROR(INDEX(ticket!B:B, MATCH(A{row_idx}, ticket!E:E, 0)), \"Not Found\")"

    # --- 3. ANALYSIS SUMMARY (for Question 2 counting) ---
    ws_analysis = wb.create_sheet(title="Analysis")
    ws_analysis.append(["Outlet ID", "Tickets Same Day", "Tickets Same Day & Same Hour"])
    for cell in ws_analysis[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    outlets = ["wrqy-juv-978", "8woh-k3u-23b"]
    for row_idx, outlet in enumerate(outlets, start=2):
        ws_analysis[f"A{row_idx}"] = outlet
        # COUNTIFS(ticket!D:D, "wrqy-juv-978", ticket!F:F, TRUE)
        ws_analysis[f"B{row_idx}"] = f"=COUNTIFS(ticket!D:D, A{row_idx}, ticket!F:F, TRUE)"
        ws_analysis[f"C{row_idx}"] = f"=COUNTIFS(ticket!D:D, A{row_idx}, ticket!G:G, TRUE)"

    # Adjust column widths
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)
    print(f"Spreadsheet generated successfully at {file_path}")

if __name__ == "__main__":
    generate_excel()
